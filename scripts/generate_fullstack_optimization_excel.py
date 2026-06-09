from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
DOCS_DIR = ROOT / "docs"
OUTPUT = DOCS_DIR / "假期旅行App全栈开发优化任务清单.xlsx"

DOCS_DIR.mkdir(parents=True, exist_ok=True)

SHEETS = {
    "优化任务清单": {
        "headers": ["任务ID", "模块", "优先级", "任务类型", "优化任务", "技术实现要点", "交互/体验要求", "数据与接口依赖", "异常/性能约束", "验收标准", "责任角色", "阶段"],
        "rows": [
            ["FS-001", "架构设计", "P0", "前后端架构", "定义Vue3前端与FastAPI后端分层架构", "前端采用Vue3 + TypeScript + Pinia + Vue-router；后端采用FastAPI异步服务；地图能力由OpenLayers统一承载", "页面模块按行程编辑、地图视图、列表排序、详情面板、导图视图拆分", "统一定义REST接口、DTO、错误码、分页与缓存策略", "禁止引入其他UI库或地图库；接口保持跨端可复用", "完成前后端模块边界、数据流和接口协议评审", "架构/前端/后端", "MVP"],
            ["FS-002", "数据模型", "P0", "数据库设计", "设计PostgreSQL核心数据模型", "使用SQLAlchemy异步ORM建立Itinerary、DayPlan、SpotNode、RouteSegment、TransportConfig、ManualLocation、RouteCache模型", "支持按天分组、节点顺序、人工定位点和路线缓存的统一管理", "数据库操作全部使用AsyncSession；建立必要索引", "节点坐标、顺序、路线段关系必须可追溯，避免重算覆盖用户手动顺序", "模型可支撑节点创建、路线计算、通行方式持久化和导图展示", "后端/DBA", "MVP"],
            ["FS-003", "前端工程", "P0", "项目骨架", "搭建Vue3 + TypeScript工程骨架", "集成Pinia、Vue-router、OpenLayers、Naive UI；配置路由、状态、API客户端和地图服务模块", "Web端与H5端共用核心业务状态，布局按端侧差异适配", "封装统一HTTP客户端与错误提示策略", "禁止混用其他UI库；组件命名和状态分层保持一致", "前端可运行并完成基础路由、状态与地图容器加载", "前端", "MVP"],
            ["FS-004", "节点管理", "P0", "业务功能", "实现景点地址输入与天数绑定", "新增节点表单必须包含景点地址/名称、所属天数；提交后调用POI解析接口生成节点", "H5端使用底部弹窗或抽屉，避免键盘遮挡核心操作", "POST /spots，POST /geo/resolve，DayPlan关联字段", "地址解析失败时允许保存文本节点或手动选点", "用户输入地址并指定天数后，可生成节点并出现在对应日期列表和地图中", "前端/后端", "MVP"],
            ["FS-005", "节点管理", "P0", "业务功能", "自动生成景点详情卡片", "详情卡包含名称、地址、所属天数、导航按钮、实景图入口、交通方案摘要", "保留地图节点点击后查看详情并唤起导航的核心链路", "GET /spots/{id}，GET /spots/{id}/resources", "POI缺失时展示用户输入内容；图片缺失不占位拉长", "任意节点可打开详情卡并触发导航入口", "前端/后端", "MVP"],
            ["FS-006", "外部资源", "P1", "资源聚合", "集成在线实景图获取与转发服务", "后端统一封装第三方图片/街景资源，前端只消费标准资源结构", "图片懒加载，详情卡首屏不因图片加载阻塞", "GET /spots/{id}/real-scenes，资源缓存字段", "弱网、无图、授权限制均需降级为基础详情展示", "有图展示、无图隐藏、弱网不影响节点详情与路线规划", "后端/前端/法务", "增强"],
            ["FS-007", "地图交互", "P0", "OpenLayers", "实现地图节点图层渲染", "使用OpenLayers VectorLayer管理节点；按天数、选中态、序号渲染样式", "点击节点打开详情卡；选中节点提升层级，路线层不得遮挡点击", "GET /spots?itineraryId=，节点坐标字段", "节点过多时启用聚合或分层渲染", "地图可展示所有节点，点击反馈准确", "前端", "MVP"],
            ["FS-008", "路线规划", "P0", "算法服务", "实现最优路线计算接口", "后端基于节点坐标、天数、通行约束计算推荐顺序和路线段；返回geometry、distance、duration", "地图上清晰绘制每日日程路线，节点序号与列表一致", "POST /routes/optimize，RouteSegment，RouteCache", "节点少于2个不计算；超时返回当前顺序与待重试状态", "同一天多个节点可生成最优路线并在地图完整呈现", "后端/前端", "MVP"],
            ["FS-009", "路线规划", "P0", "地图可视化", "实现OpenLayers路径连线与路线层管理", "按Day拆分路线VectorLayer；当前日高亮，非当前日淡化或隐藏", "支持整体行程查看和单日查看切换", "GET /routes?itineraryId=&day=，路线geometry", "避免一次性绘制过多复杂线段造成前端卡顿", "切换日期后地图路线层正确刷新且不卡顿", "前端", "MVP"],
            ["FS-010", "拖拽排序", "P0", "业务功能", "实现独立节点排序列表视图", "使用Naive UI列表与拖拽交互封装排序面板；拖拽后更新节点order并触发路线重算", "拖拽时显示目标插入位置、序号变化和重算状态", "PATCH /spots/reorder，POST /routes/recalculate", "重算失败保留用户手动顺序，不回滚用户操作", "拖拽后列表顺序、地图编号和路线连线同步更新", "前端/后端", "MVP"],
            ["FS-011", "路线状态", "P0", "状态管理", "区分系统推荐路线与用户手动路线", "Pinia保存routeSource、sortMode、routeStatus；后端持久化排序来源与更新时间", "界面展示“系统推荐 / 已手动调整 / 待重算”标签和重置入口", "SortState、RouteCache、PATCH /routes/reset", "重置推荐不应误删用户节点数据", "用户可识别路线来源并恢复系统推荐", "前端/后端", "MVP"],
            ["FS-012", "通行方式", "P0", "业务功能", "实现节点间通行方式配置", "支持步行、自驾、公交、自定义；路线段与交通配置分离存储", "点击路线段或详情交通模块打开配置面板", "POST /transport-configs，GET /route-segments/{id}", "SDK或算法不支持时允许自定义输入耗时/距离", "任意两节点间可保存通行方式并在详情页展示", "前端/后端", "MVP"],
            ["FS-013", "人工定位", "P1", "业务功能", "支持节点与人工定位地址之间配置通行方式", "OpenLayers地图选点生成ManualLocation；可作为临时端点参与交通方案配置", "人工点使用不同图标或虚线标识，避免误认为正式景点", "POST /manual-locations，POST /transport-configs", "定位未授权时允许手动地图选点；人工点可设置有效期", "人工定位点可参与通行方式配置且可持久化", "前端/后端", "增强"],
            ["FS-014", "思维导图", "P1", "业务功能", "实现行程思维导图视图", "前端从统一行程数据派生导图节点；按Day为根节点，景点为子节点，路线段为关联边", "概览区域支持地图/导图切换，点击导图节点可定位地图或打开详情", "GET /itineraries/{id}/mindmap 或前端派生结构", "导图不得维护独立不可同步数据源", "导图顺序、天数、节点详情与地图/列表保持一致", "前端/后端", "增强"],
            ["FS-015", "布局优化", "P0", "交互体验", "重构地图与详情面板纵向空间", "地图容器与详情面板采用折叠/悬浮/分屏策略；H5优先三段式底部面板，Web可使用侧栏分屏", "详情摘要态仅保留核心信息和导航按钮；完整信息进入展开态", "前端布局状态与路由查询参数可选同步", "不得形成“地图+地址详情”纵向长串阻塞滑动", "页面上下滑动顺畅，地图区域仍保留足够可操作空间", "前端/UI", "MVP"],
            ["FS-016", "滑动隔离", "P0", "交互体验", "解决地图操作与页面滚动冲突", "明确mapDrag、panelDrag、listScroll状态；在不同区域绑定独立事件策略", "地图拖拽、缩放与页面垂直滚动互不抢占；面板手柄专门用于高度切换", "前端手势状态机与组件事件协议", "低端机降低动画复杂度，避免滚动掉帧", "H5端上下滑动浏览不再误触地图拖拽", "前端/UI", "MVP"],
            ["FS-017", "跨端适配", "P0", "响应式", "完成Web端与H5端布局适配", "Web端可采用地图+侧栏+详情抽屉；H5端采用地图+底部面板+列表切换", "按钮、列表项、拖拽手柄满足移动触控热区规范", "CSS变量、断点配置、端侧识别", "安全区、键盘、横竖屏需要适配", "主流浏览器与H5小屏设备无遮挡、无溢出、无操作死区", "前端/UI", "MVP"],
            ["FS-018", "性能优化", "P0", "性能", "优化节点数量增加时的地图与路线性能", "节点渲染分层、路线缓存、图片懒加载、必要时聚合节点；路线算法结果缓存", "加载状态轻量展示，不阻塞核心编辑流程", "RouteCache、资源缓存、前端虚拟列表", "避免重复请求、重复绘制和大图首屏加载", "节点数量增加时地图缩放、拖拽、路线切换仍保持流畅", "前端/后端", "MVP"],
            ["FS-019", "后端接口", "P0", "API开发", "实现景点数据CRUD接口", "FastAPI异步接口支持创建、查询、编辑、删除、跨天移动、排序更新", "前端所有节点操作使用统一接口反馈状态", "POST/GET/PATCH/DELETE /spots", "删除节点需同步处理路线段与交通配置关系", "节点增删改查稳定，错误码清晰", "后端", "MVP"],
            ["FS-020", "后端接口", "P0", "API开发", "实现路线规划与重算接口", "接口接收节点集合、排序策略、交通约束，返回路线段和路线状态", "前端可展示计算中、成功、失败、待重试状态", "POST /routes/optimize，POST /routes/recalculate", "规划超时保留用户当前顺序，返回可降级状态", "路线计算接口可支撑自动生成和拖拽后重算", "后端", "MVP"],
            ["FS-021", "后端接口", "P1", "API开发", "实现通行方式存储接口", "支持节点-节点、节点-人工点两类端点；交通方式字典可扩展", "详情页按路线段展示完整交通方案", "POST/GET/PATCH /transport-configs", "未知交通类型按custom处理，不导致前端崩溃", "交通方式可持久化、可编辑、可展示", "后端/前端", "增强"],
            ["FS-022", "资源服务", "P1", "API开发", "实现实景图在线获取与文件资源服务", "后端转发或缓存第三方资源，统一输出图片URL、来源、授权状态、失效时间", "前端懒加载，失败后展示无图态", "GET /resources/real-scenes，文件服务路由", "遵守第三方平台授权与缓存规则", "资源服务失败不影响节点详情和路线主链路", "后端/法务", "增强"],
            ["FS-023", "状态管理", "P0", "前端状态", "设计Pinia行程状态模型", "按itinerary、dayPlans、spots、routes、transportConfigs、uiState拆分store", "状态变化驱动地图、列表、详情和导图同步更新", "前端API响应结构统一映射", "避免导图、地图、列表各自维护不一致副本", "任一节点顺序变化后四类视图保持一致", "前端", "MVP"],
            ["FS-024", "权限与合规", "P1", "合规", "处理定位、地图服务与图片资源授权", "定位前说明用途；地图与图片资源记录供应商与授权边界", "拒绝定位后允许手动选点继续规划", "权限状态、资源来源字段", "不得强制授权定位；不得绕过地图平台授权", "拒绝授权、无图、资源受限时都有可用替代路径", "产品/法务/前端/后端", "增强"],
            ["FS-025", "测试验收", "P0", "质量保障", "建立端到端验收用例", "覆盖节点创建、POI解析、路线生成、拖拽排序、通行方式、导图切换、滑动隔离、弱网降级", "用真实H5尺寸验证触控热区和滑动体验", "测试数据、接口Mock、弱网模拟", "区分功能失败、接口失败、地图SDK限制和权限失败", "P0链路均可按用户旅程稳定通过", "测试/前端/后端", "MVP"],
        ],
    },
    "架构拆解": {
        "headers": ["层级", "模块", "职责", "关键依赖", "输出物"],
        "rows": [
            ["前端", "Vue3应用层", "承载行程编辑、地图、排序、详情、导图和适配逻辑", "Vue3、TypeScript、Pinia、Vue-router、Naive UI", "页面、组件、Store、路由"],
            ["前端", "OpenLayers地图层", "节点渲染、路线连线、人工选点、地图点击与图层管理", "OpenLayers VectorLayer、Feature、Geometry", "节点层、路线层、选中态、人工点层"],
            ["后端", "FastAPI服务层", "提供景点、路线、交通配置、资源聚合接口", "FastAPI、Pydantic、异步任务", "REST API、错误码、接口文档"],
            ["后端", "算法与缓存层", "计算推荐路线、缓存路线结果、支持拖拽后重算", "路线算法、RouteCache、PostgreSQL", "路线段、geometry、distance、duration"],
            ["数据", "PostgreSQL持久层", "保存行程、节点、天数、排序、交通方式、人工定位点", "SQLAlchemy AsyncSession", "结构化业务数据"],
            ["资源", "第三方资源聚合", "获取并转发实景图、街景或POI资源", "地图/图片平台授权接口", "标准化资源URL与来源信息"],
        ],
    },
    "前端模块": {
        "headers": ["模块", "技术栈", "核心组件/状态", "交互重点", "风险控制"],
        "rows": [
            ["行程编辑", "Vue3 + Naive UI", "SpotForm、DaySelector、useItineraryStore", "输入地址必须绑定天数，支持手动选点兜底", "键盘遮挡、重复提交、地址解析失败"],
            ["地图视图", "OpenLayers", "MapCanvas、SpotLayer、RouteLayer、ManualPointLayer", "节点点击、路线高亮、人工选点", "图层遮挡、过量Feature、地图事件抢占滚动"],
            ["排序列表", "Vue3 + Pinia", "SortableSpotList、RouteStatusBar", "长按拖拽、序号反馈、重算状态", "重算失败不回滚用户顺序"],
            ["详情面板", "Naive UI + 响应式CSS", "SpotDetailSheet、TransportSummary、RealScenePreview", "摘要态/半屏态/全屏态切换", "纵向长串、左右过满、图片加载阻塞"],
            ["通行方式", "Naive UI", "TransportConfigPanel、RouteSegmentCard", "路线段选择、方式配置、自定义耗时距离", "交通字典扩展、SDK不支持降级"],
            ["思维导图", "Vue3组件化", "MindMapView、DayRootNode、SpotMindNode", "地图/导图切换、节点联动详情", "不能维护独立数据副本"],
        ],
    },
    "后端与数据模型": {
        "headers": ["模型/服务", "关键字段/能力", "用途", "注意事项"],
        "rows": [
            ["Itinerary", "id, title, start_date, view_mode", "行程主对象", "视图偏好可扩展"],
            ["DayPlan", "id, itinerary_id, day_index, date", "按天组织节点", "day_index与节点order分离"],
            ["SpotNode", "id, day_plan_id, name, address, poi_id, lat, lng, order_index", "景点节点", "支持POI节点、文本节点、手动点补全"],
            ["RouteSegment", "id, from_node_id, to_node_id, geometry, distance, duration, status", "节点间路线段", "geometry可为空，支持失败态"],
            ["TransportConfig", "id, from_ref, to_ref, mode, custom_note, duration, distance", "通行方式配置", "端点可为节点或人工定位点"],
            ["ManualLocation", "id, label, lat, lng, source, expires_at", "人工地图选点", "拒绝定位时仍可手动创建"],
            ["RouteCache", "id, itinerary_id, day_index, input_hash, result_json, expires_at", "路线计算缓存", "避免重复计算和重复绘制"],
            ["ResourceProxy", "spot_id, url, provider, license_status, expires_at", "实景图与街景资源", "遵守第三方授权和缓存周期"],
        ],
    },
    "API接口清单": {
        "headers": ["接口", "方法", "用途", "输入重点", "输出重点", "异常策略"],
        "rows": [
            ["/spots", "POST", "创建景点节点", "address/name/dayIndex", "SpotNode", "地址解析失败允许保存文本节点"],
            ["/spots", "GET", "查询节点列表", "itineraryId/dayIndex", "按天分组节点", "无数据返回空列表"],
            ["/spots/{id}", "PATCH", "编辑节点", "name/address/dayIndex/order", "更新后节点", "跨天移动触发路线状态变更"],
            ["/spots/reorder", "PATCH", "拖拽排序", "dayIndex + orderedNodeIds", "SortState", "失败时返回原顺序与错误原因"],
            ["/geo/resolve", "POST", "地址解析", "keyword/address", "POI候选与坐标", "低置信度返回候选而非强制失败"],
            ["/routes/optimize", "POST", "最优路线计算", "nodeIds/transportConstraints", "RouteSegment[]", "超时返回待重试状态"],
            ["/routes/recalculate", "POST", "排序后重算", "orderedNodeIds/sortSource", "新路线段与状态", "保留用户排序"],
            ["/transport-configs", "POST", "保存通行方式", "fromRef/toRef/mode", "TransportConfig", "不支持交通方式转自定义"],
            ["/manual-locations", "POST", "创建人工定位点", "lat/lng/label/source", "ManualLocation", "定位拒绝不影响手动选点"],
            ["/resources/real-scenes", "GET", "获取实景资源", "spotId/poiId", "图片URL与来源", "无图返回空资源列表"],
        ],
    },
    "地图交互与布局": {
        "headers": ["场景", "推荐实现", "交互规则", "验收关注点"],
        "rows": [
            ["地图节点点击", "OpenLayers Feature绑定节点ID并统一处理点击事件", "点击节点打开详情摘要卡，保留导航入口", "路线层不遮挡节点点击"],
            ["路线绘制", "按Day创建路线VectorLayer，当前日高亮", "整体/单日切换时刷新图层", "切换流畅且序号一致"],
            ["人工选点", "进入选点模式后点击地图生成ManualLocation", "人工点与正式景点视觉区分", "可参与交通方式配置"],
            ["H5详情面板", "三段式底部面板：摘要态、半屏态、全屏态", "手柄拖动控制高度，内容区独立滚动", "不阻塞页面上下滑动"],
            ["Web布局", "地图主视区 + 右侧列表/详情栏", "大屏展示更多信息，小屏自动切底部面板", "无左右过满、无按钮遮挡"],
            ["滑动隔离", "mapDrag/panelDrag/listScroll状态机", "地图区域响应地图手势，面板内容响应滚动", "H5端滑动不误触地图"],
        ],
    },
    "异常性能合规": {
        "headers": ["类别", "场景", "处理策略", "用户提示", "是否阻断"],
        "rows": [
            ["异常", "地址解析失败", "允许保存文本节点或手动地图选点", "未找到准确地点，可手动选择位置", "否"],
            ["异常", "路线规划超时", "保留当前顺序，显示虚线或待计算状态", "路线计算较慢，已先按当前顺序展示", "否"],
            ["异常", "实景图缺失", "隐藏图片区，保留详情和导航", "暂无实景图，不影响行程规划", "否"],
            ["异常", "定位未授权", "允许手动选点", "未开启定位，可在地图上手动选择地址", "否"],
            ["性能", "节点数量增加", "节点聚合、分层绘制、路线缓存、虚拟列表", "已优化当前设备展示", "否"],
            ["性能", "图片资源较多", "懒加载、占位骨架、失败重试", "图片加载中", "否"],
            ["合规", "地图/图片授权", "记录供应商、用途、缓存周期，遵守平台规范", "资源受限时展示基础信息", "否"],
            ["合规", "定位隐私", "先说明用途，再请求权限，支持拒绝后继续规划", "定位仅用于路线规划和选点", "否"],
        ],
    },
    "验收路线图": {
        "headers": ["阶段", "范围", "关键交付", "验收标准"],
        "rows": [
            ["MVP", "P0核心链路", "前后端骨架、节点管理、地图节点、最优路线、拖拽排序、详情面板、滑动隔离", "可完成创建节点→生成路线→拖拽重算→查看详情→导航的闭环"],
            ["增强", "P1体验扩展", "实景图聚合、人工定位点、通行方式扩展、思维导图、合规完善", "功能可用且不影响MVP主链路"],
            ["优化", "性能与适配", "路线缓存、节点聚合、图片懒加载、Web/H5断点优化", "多节点、多图片、小屏设备下操作流畅"],
            ["评审", "产品与研发验收", "接口文档、数据模型、交互验收清单、异常用例", "产品、前端、后端、测试可按表分工落地"],
        ],
    },
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BORDER_SIDE = Side(style="thin", color="D9E2F3")
CELL_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)
PRIORITY_FILL = {
    "P0": PatternFill("solid", fgColor="FCE4D6"),
    "P1": PatternFill("solid", fgColor="E2F0D9"),
    "MVP": PatternFill("solid", fgColor="DDEBF7"),
    "增强": PatternFill("solid", fgColor="E2F0D9"),
    "优化": PatternFill("solid", fgColor="FFF2CC"),
}


def append_sheet(wb: Workbook, title: str, payload: dict, index: int):
    ws = wb.active if index == 1 else wb.create_sheet(title)
    ws.title = title
    ws.append(payload["headers"])
    for row in payload["rows"]:
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = CELL_BORDER
            if cell.value in PRIORITY_FILL:
                cell.fill = PRIORITY_FILL[cell.value]

    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        values = [str(ws.cell(row=row_idx, column=col_idx).value or "") for row_idx in range(1, ws.max_row + 1)]
        width = min(max(max(len(value) for value in values) + 2, 10), 42)
        if title == "优化任务清单":
            width_map = {1: 10, 2: 12, 3: 10, 4: 14, 5: 30, 6: 38, 7: 36, 8: 30, 9: 34, 10: 34, 11: 18, 12: 12}
            width = width_map.get(col_idx, width)
        ws.column_dimensions[letter].width = width

    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 28 if row_idx == 1 else (86 if title == "优化任务清单" else 60)

    table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=f"TaskTable{index}", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def main():
    wb = Workbook()
    for index, (title, payload) in enumerate(SHEETS.items(), start=1):
        append_sheet(wb, title, payload, index)
    wb.save(OUTPUT)
    print(f"generated: {OUTPUT}")


if __name__ == "__main__":
    main()